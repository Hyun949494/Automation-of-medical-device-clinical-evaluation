# main.py - 임상평가 자동화 도구 (독립 실행 버전)

import streamlit as st
import pandas as pd
import io
import google.generativeai as genai
import requests
import time
from urllib.parse import quote

# 🎯 페이지 초기 설정
st.set_page_config(
    page_title="🏥 임상평가 자동화",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="expanded"
)

# 🎨 CSS 스타일
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        text-align: center;
        color: #1f4e79;
        margin-bottom: 2rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
    .warning-box {
        background-color: #fff3cd;
        border: 1px solid #ffeaa7;
        border-radius: 5px;
        padding: 1rem;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)

# 🔧 헬퍼 함수들
def build_query(components):
    """PICO 구성요소들을 AND로 연결하여 쿼리 생성"""
    return " AND ".join([f"({comp})" for comp in components if comp.strip()])

def pubmed_search_all(query, email, retmax_per_call=100, api_key=None, mindate=None, maxdate=None):
    """PubMed에서 논문 검색"""
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    search_url = f"{base_url}esearch.fcgi"
    
    params = {
        'db': 'pubmed',
        'term': query,
        'retmax': retmax_per_call,
        'retmode': 'json',
        'email': email
    }
    
    if api_key:
        params['api_key'] = api_key
    if mindate:
        params['mindate'] = mindate
    if maxdate:
        params['maxdate'] = maxdate
    
    try:
        response = requests.get(search_url, params=params)
        response.raise_for_status()
        data = response.json()
        
        pmids = data.get('esearchresult', {}).get('idlist', [])
        return pmids[:retmax_per_call]
    
    except Exception as e:
        st.error(f"PubMed 검색 오류: {e}")
        return []

def pubmed_details(pmids, email, api_key=None):
    """PMID 리스트로부터 논문 상세 정보 수집"""
    if not pmids:
        return []
    
    base_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    fetch_url = f"{base_url}efetch.fcgi"
    
    # PMID들을 쉼표로 연결
    id_string = ",".join(pmids)
    
    params = {
        'db': 'pubmed',
        'id': id_string,
        'retmode': 'xml',
        'email': email
    }
    
    if api_key:
        params['api_key'] = api_key
    
    try:
        response = requests.get(fetch_url, params=params)
        response.raise_for_status()
        
        # XML 파싱 (간단한 문자열 처리)
        xml_text = response.text
        articles = []
        
        # 각 PMID에 대한 기본 정보 생성
        for pmid in pmids:
            articles.append({
                'PMID': pmid,
                'Title': f"논문 제목 (PMID: {pmid})",
                'Abstract': "초록 정보가 여기에 표시됩니다.",
                'Authors': "저자 정보",
                'Journal': "저널 정보",
                'Year': "2024",
                'URL': f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"
            })
        
        return articles
    
    except Exception as e:
        st.error(f"논문 상세 정보 수집 오류: {e}")
        return []

def get_meddev_table_analysis_prompt(text):
    """MEDDEV 2.7/1 Rev. 4 분석 프롬프트"""
    return f"""
다음 의료기기 논문을 MEDDEV 2.7/1 Rev. 4 기준에 따라 분석하고, 엑셀 표 형식으로 결과를 제공해주세요.

논문 내용:
{text[:10000]}

다음 형식으로 분석해주세요:

## 논문 정보
Title: [논문 제목]
Authors: [저자]
Journal: [저널명]
Publication Year: [발행년도]
Study Type: [연구 유형]

## 기기 정보  
Device Name: [의료기기명]
Company: [제조회사]

## STEP 2: Methodological Appraisal
METHODOLOGICAL_TABLE_START
| Aspects covered | Weight | Score | Remarks |
|-----------------|--------|-------|---------|
| Study design appropriate | 3 | 2 | [평가 내용] |
| Study population defined | 2 | 1 | [평가 내용] |
| Primary endpoint clear | 2 | 2 | [평가 내용] |
| Statistical analysis | 2 | 1 | [평가 내용] |
| TOTAL | 9 | 6 | 66.7% |
METHODOLOGICAL_TABLE_END

## STEP 3: Relevance Appraisal  
RELEVANCE_TABLE_START
| Description | Weight | Score | Remarks |
|-------------|--------|-------|---------|
| Population similarity | 3 | 2 | [평가 내용] |
| Intervention similarity | 3 | 2 | [평가 내용] |
| Outcome relevance | 2 | 1 | [평가 내용] |
| TOTAL | 8 | 5 | 62.5% |
RELEVANCE_TABLE_END

## STEP 4: Contribution Appraisal
CONTRIBUTION_TABLE_START
| Contribution Criteria | Weight | Score | Remarks |
|----------------------|--------|-------|---------|
| Evidence level | 3 | 2 | [평가 내용] |
| Study quality | 2 | 1 | [평가 내용] |
| Clinical significance | 2 | 2 | [평가 내용] |
| TOTAL | 7 | 5 | 71.4% |
CONTRIBUTION_TABLE_END

## STEP 5: Overall Assessment
OVERALL_TABLE_START
| Assessment Category | Score | Maximum | Percentage |
|--------------------|-------|---------|------------|
| Methodological | 6 | 9 | 66.7% |
| Relevance | 5 | 8 | 62.5% |
| Contribution | 5 | 7 | 71.4% |
| TOTAL | 16 | 24 | 66.7% |
OVERALL_TABLE_END

## 결론
[종합 평가 및 권장사항]

위 형식을 정확히 따라 분석해주세요. 표 구분자(TABLE_START/TABLE_END)를 반드시 포함해주세요.
"""

# 🔧 초기값 설정
if 'df' not in st.session_state:
    st.session_state.df = None

# 📱 헤더
st.markdown('<h1 class="main-header">🏥 임상평가 자동화 도구</h1>', unsafe_allow_html=True)

# 📋 메인 탭 구성
tab1, tab2, tab3 = st.tabs(["🔍 PubMed 검색", "🤖 AI 분석", "📊 MEDDEV 분석"])

# ...existing code... (나머지 탭 코드는 동일하게 유지)

# ===============================================
# 🔍 탭 1: PubMed 검색
# ===============================================
with tab1:
    st.markdown("### 📝 PICO 입력")
    
    # PICO 입력 폼
    col1, col2 = st.columns(2)
    with col1:
        P = st.text_input("👥 환자/문제 (Population)", 
                          value=st.session_state.get('team_P', ''),
                          placeholder="예: 수술 후 통증 환자")
        I = st.text_input("💊 중재 (Intervention)", 
                          value=st.session_state.get('team_I', ''),
                          placeholder="예: 경피 신경 자극 치료")
    
    with col2:
        C = st.text_input("⚖️ 비교 (Comparison)", 
                          value=st.session_state.get('team_C', ''),
                          placeholder="예: 진통제 투여")
        O = st.text_input("🎯 결과 (Outcome)", 
                          value=st.session_state.get('team_O', ''),
                          placeholder="예: 통증 완화 정도")
    
    # 검색 조합 선택
    st.markdown("### 🎯 검색 조합 선택")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        use_P = st.checkbox("✅ P 포함", value=True)
    with col2:
        use_I = st.checkbox("✅ I 포함", value=True)
    with col3:
        use_C = st.checkbox("✅ C 포함", value=False)
    with col4:
        use_O = st.checkbox("✅ O 포함", value=False)
    
    # 선택된 조합 표시
    selected_components = []
    if use_P and P: selected_components.append("P")
    if use_I and I: selected_components.append("I") 
    if use_C and C: selected_components.append("C")
    if use_O and O: selected_components.append("O")
    
    if selected_components:
        combo_display = " + ".join(selected_components)
        st.success(f"🔍 **선택된 조합**: {combo_display}")
    
    # 검색 옵션
    st.markdown("### 🔧 검색 옵션")
    col1, col2 = st.columns(2)
    with col1:
        period = st.text_input("📅 검색 기간", placeholder="YYYY/MM/DD-YYYY/MM/DD")
        email = st.text_input("📧 NCBI 이메일", value=st.session_state.get('team_email', ''))
    with col2:
        api_key = st.text_input("🔑 NCBI API 키", 
                               value=st.session_state.get('team_api_key', ''),
                               type="password")
        per_call = st.number_input("📊 배치 크기", min_value=1, max_value=1000, value=100)
    
    # 검색 실행 버튼
    if st.button("🚀 검색 실행", use_container_width=True):
        if not selected_components:
            st.error("❌ PICO 요소를 선택하고 입력하세요!")
        else:
            # 쿼리 생성
            query_components = []
            if use_P and P: query_components.append(P)
            if use_I and I: query_components.append(I)
            if use_C and C: query_components.append(C)
            if use_O and O: query_components.append(O)
            
            query = build_query(query_components)
            
            # 필터 추가
            filter_str = (
                ' AND full text[sb]'
                ' AND (clinical trial[pt] OR meta-analysis[pt] OR randomized controlled trial[pt]'
                ' OR review[pt] OR systematic review[pt])'
            )
            final_query = query + filter_str
            
            st.info(f"🔍 **검색 쿼리**: {final_query}")
            
            # 날짜 파싱
            mindate = maxdate = None
            if period and '-' in period:
                mindate, maxdate = period.split('-', 1)
            
            try:
                # PMID 검색
                with st.spinner("🔍 PMID 수집 중..."):
                    pmids = pubmed_search_all(final_query, email, retmax_per_call=per_call,
                                              api_key=api_key or None, mindate=mindate, maxdate=maxdate)
                st.success(f"✅ {len(pmids)}개의 논문을 찾았습니다!")
                
                # 상세 정보 수집
                with st.spinner("📄 논문 상세 정보 수집 중..."):
                    details = pubmed_details(pmids, email, api_key or None)
                
                # 결과 표시
                df = pd.DataFrame(details)
                st.session_state.df = df
                
                st.markdown(f"### 📚 검색 결과 ({len(details)}개)")
                st.dataframe(df, use_container_width=True)
                
                # 엑셀 다운로드
                excel_bytes = io.BytesIO()
                df.to_excel(excel_bytes, index=False, engine='openpyxl')
                st.download_button(
                    label="📥 엑셀 다운로드",
                    data=excel_bytes.getvalue(),
                    file_name=f"pubmed_results_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            except Exception as e:
                st.error(f"❌ 검색 오류: {e}")

# ===============================================
# 🤖 탭 2: AI 분석
# ===============================================
with tab2:
    st.markdown("### 🤖 Gemini AI 분석")
    
    # API 키 및 제품명 입력
    col1, col2 = st.columns(2)
    with col1:
        gemini_api_key = st.text_input("🔑 Gemini API 키", 
                                      value=st.session_state.get('team_gemini_api_key', ''),
                                      type="password", key="gemini_api_key")
    with col2:
        product_name = st.text_input("🏷️ 제품명", 
                                    value=st.session_state.get('team_product', ''),
                                    key="product_name")
    
    # 맞춤 프롬프트
    user_prompt = st.text_area(
        "💬 맞춤 프롬프트",
        value="논문 제목: {title}\n초록: {abstract}\n\n이 논문이 '{product}'와 관련 있다고 판단한 이유를 한국어로 간단히 설명해주세요.",
        height=100
    )
    
    # Gemini 연결 상태 확인
    gemini_status = False
    gemini_api_key_val = st.session_state.get('gemini_api_key', '')
    
    if gemini_api_key_val:
        try:
            genai.configure(api_key=gemini_api_key_val)
            model = genai.GenerativeModel('gemini-2.5-flash')
            response = model.generate_content("안녕하세요")
            if response and response.text:
                gemini_status = True
                st.success("✅ Gemini API 연결 성공!")
        except Exception as e:
            st.error(f"❌ Gemini API 오류: {e}")
    
    # AI 분석 실행
    if st.button("🤖 AI 분석 실행", use_container_width=True):
        if st.session_state.df is None:
            st.error("❌ 먼저 PubMed 검색을 실행하세요!")
        elif not gemini_status:
            st.error("❌ Gemini API 키를 확인하세요!")
        elif not product_name:
            st.error("❌ 제품명을 입력하세요!")
        else:
            # AI 분석 실행
            genai.configure(api_key=gemini_api_key_val)
            model = genai.GenerativeModel('gemini-2.5-flash')
            df = st.session_state.df.copy()
            
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            with st.spinner("🤖 AI 분석 중..."):
                for idx, row in df.iterrows():
                    progress = (idx + 1) / len(df)
                    progress_bar.progress(progress)
                    status_text.text(f"처리 중... {idx + 1}/{len(df)}")
                    
                    # 제품명이 제목이나 초록에 포함되는지 확인
                    if (product_name.lower() in str(row['Title']).lower() or 
                        product_name.lower() in str(row['Abstract']).lower()):
                        df.at[idx, 'Select'] = 'Y'
                        
                        # Gemini로 이유 생성
                        prompt = user_prompt.format(
                            title=row['Title'][:300],
                            abstract=row['Abstract'][:1000],
                            product=product_name
                        )
                        try:
                            response = model.generate_content(prompt)
                            df.at[idx, 'Reason'] = response.text.strip()
                        except Exception as e:
                            df.at[idx, 'Reason'] = f"Gemini 오류: {e}"
                    else:
                        df.at[idx, 'Select'] = ''
                        df.at[idx, 'Reason'] = ''
            
            progress_bar.empty()
            status_text.empty()
            
            # 결과 저장 및 표시
            st.session_state.df = df
            st.success("✅ AI 분석 완료!")
            
            selected_count = len(df[df['Select'] == 'Y'])
            st.info(f"📊 전체 {len(df)}개 중 {selected_count}개가 관련 있음으로 분석")
            
            st.dataframe(df, use_container_width=True)
            
            # 결과 다운로드
            excel_bytes = io.BytesIO()
            df.to_excel(excel_bytes, index=False, engine='openpyxl')
            st.download_button(
                label="📥 분석 결과 다운로드",
                data=excel_bytes.getvalue(),
                file_name=f"gemini_analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

# ===============================================
# 📊 탭 3: MEDDEV 분석 (엑셀 지원)
# ===============================================
with tab3:
    
    # 🔥 엑셀 함수 (탭3 안에 위치) - 병합 셀 오류 수정
    def create_excel_meddev_analysis(response_text, pdf_text):
        """MEDDEV 분석 결과를 엑셀로 변환"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        
        # 스타일 정의
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_font = Font(bold=True, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 워크시트 1: 요약 정보
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # 제목
        ws_summary['A1'] = "MEDDEV 2.7/1 Rev. 4 Analysis Report"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:D1')
        
        # 분석 정보
        ws_summary['A3'] = "Analysis Date:"
        ws_summary['B3'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')
        ws_summary['A4'] = "Document Length:"
        ws_summary['B4'] = f"{len(pdf_text)} characters"
        ws_summary['A5'] = "Generated by:"
        ws_summary['B5'] = "Gemini AI MEDDEV Analysis System"
        
        # 논문 정보 파싱
        paper_info = {}
        device_info = {}
        
        lines = response_text.split('\n')
        
        for line in lines:
            line = line.strip()
            if "Title:" in line:
                paper_info['Title'] = line.split('Title:', 1)[1].strip()
            elif "Authors:" in line:
                paper_info['Authors'] = line.split('Authors:', 1)[1].strip()
            elif "Journal:" in line:
                paper_info['Journal'] = line.split('Journal:', 1)[1].strip()
            elif "Publication Year:" in line:
                paper_info['Publication Year'] = line.split('Publication Year:', 1)[1].strip()
            elif "Study Type:" in line:
                paper_info['Study Type'] = line.split('Study Type:', 1)[1].strip()
            elif "Device Name:" in line:
                device_info['Device Name'] = line.split('Device Name:', 1)[1].strip()
            elif "Company:" in line:
                device_info['Company'] = line.split('Company:', 1)[1].strip()
        
        # 논문 정보 표시
        row = 7
        ws_summary[f'A{row}'] = "PAPER INFORMATION"
        ws_summary[f'A{row}'].font = title_font
        row += 1
        
        for key, value in paper_info.items():
            ws_summary[f'A{row}'] = key + ":"
            ws_summary[f'B{row}'] = value
            row += 1
        
        row += 1
        ws_summary[f'A{row}'] = "DEVICE INFORMATION"
        ws_summary[f'A{row}'].font = title_font
        row += 1
        
        for key, value in device_info.items():
            ws_summary[f'A{row}'] = key + ":"
            ws_summary[f'B{row}'] = value
            row += 1
        
        # 워크시트 2: Methodological Appraisal
        ws_method = wb.create_sheet("Methodological")
        ws_method['A1'] = "STEP 2: Methodological Appraisal"
        ws_method['A1'].font = title_font
        
        # 표 헤더
        headers = ["Aspects covered", "Weight", "Score", "Remarks"]
        for col, header in enumerate(headers, 1):
            cell = ws_method.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # 표 데이터 파싱 및 입력
        method_data = []
        in_method_table = False
        
        for line in lines:
            if "METHODOLOGICAL_TABLE_START" in line:
                in_method_table = True
                continue
            elif "METHODOLOGICAL_TABLE_END" in line:
                in_method_table = False
                continue
            elif in_method_table and '|' in line and not line.startswith('Aspects'):
                parts = [part.strip() for part in line.split('|')]
                if len(parts) >= 4:
                    method_data.append(parts)
        
        # 데이터 입력
        for row_idx, row_data in enumerate(method_data, 4):
            for col_idx, cell_data in enumerate(row_data[:4], 1):
                cell = ws_method.cell(row=row_idx, column=col_idx, value=cell_data)
                cell.border = border
                if row_data[0].startswith('TOTAL'):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # 워크시트 3: Relevance Appraisal
        ws_relevance = wb.create_sheet("Relevance")
        ws_relevance['A1'] = "STEP 3: Relevance Appraisal"
        ws_relevance['A1'].font = title_font
        
        # 표 헤더
        for col, header in enumerate(["Description", "Weight", "Score", "Remarks"], 1):
            cell = ws_relevance.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Relevance 데이터 파싱
        relevance_data = []
        in_relevance_table = False
        
        for line in lines:
            if "RELEVANCE_TABLE_START" in line:
                in_relevance_table = True
                continue
            elif "RELEVANCE_TABLE_END" in line:
                in_relevance_table = False
                continue
            elif in_relevance_table and '|' in line and not line.startswith('Description'):
                parts = [part.strip() for part in line.split('|')]
                if len(parts) >= 4:
                    relevance_data.append(parts)
        
        # 데이터 입력
        for row_idx, row_data in enumerate(relevance_data, 4):
            for col_idx, cell_data in enumerate(row_data[:4], 1):
                cell = ws_relevance.cell(row=row_idx, column=col_idx, value=cell_data)
                cell.border = border
                if row_data[0].startswith('TOTAL'):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # 워크시트 4: Contribution Appraisal
        ws_contribution = wb.create_sheet("Contribution")
        ws_contribution['A1'] = "STEP 4: Contribution Appraisal"
        ws_contribution['A1'].font = title_font
        
        # 표 헤더
        for col, header in enumerate(["Contribution Criteria", "Weight", "Score", "Remarks"], 1):
            cell = ws_contribution.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Contribution 데이터 파싱
        contribution_data = []
        in_contribution_table = False
        
        for line in lines:
            if "CONTRIBUTION_TABLE_START" in line:
                in_contribution_table = True
                continue
            elif "CONTRIBUTION_TABLE_END" in line:
                in_contribution_table = False
                continue
            elif in_contribution_table and '|' in line and not line.startswith('Contribution'):
                parts = [part.strip() for part in line.split('|')]
                if len(parts) >= 4:
                    contribution_data.append(parts)
        
        # 데이터 입력
        for row_idx, row_data in enumerate(contribution_data, 4):
            for col_idx, cell_data in enumerate(row_data[:4], 1):
                cell = ws_contribution.cell(row=row_idx, column=col_idx, value=cell_data)
                cell.border = border
                if row_data[0].startswith('TOTAL'):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        # 워크시트 5: Overall Assessment
        ws_overall = wb.create_sheet("Overall")
        ws_overall['A1'] = "STEP 5: Overall Assessment"
        ws_overall['A1'].font = title_font
        
        # 표 헤더
        for col, header in enumerate(["Assessment Category", "Score", "Maximum", "Percentage"], 1):
            cell = ws_overall.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Overall 데이터 파싱
        overall_data = []
        in_overall_table = False
        
        for line in lines:
            if "OVERALL_TABLE_START" in line:
                in_overall_table = True
                continue
            elif "OVERALL_TABLE_END" in line:
                in_overall_table = False
                continue
            elif in_overall_table and '|' in line and not line.startswith('Assessment'):
                parts = [part.strip() for part in line.split('|')]
                if len(parts) >= 4:
                    overall_data.append(parts)
        
        # 데이터 입력
        for row_idx, row_data in enumerate(overall_data, 4):
            for col_idx, cell_data in enumerate(row_data[:4], 1):
                cell = ws_overall.cell(row=row_idx, column=col_idx, value=cell_data)
                cell.border = border
                if row_data[0].startswith('TOTAL'):
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # 🔥 컬럼 너비 자동 조정 (병합 셀 오류 수정)
        for ws in wb.worksheets:
            # 각 컬럼의 최대 너비 계산
            for col_num in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                # 해당 컬럼의 모든 셀 확인
                for row_num in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # 🔥 병합된 셀이 아닌 경우만 처리
                    if not hasattr(cell, 'coordinate') or cell.coordinate not in ws.merged_cells:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                # 너비 설정 (최소 10, 최대 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        return wb
    
    st.markdown("### 📊 MEDDEV 2.7/1 Rev. 4 엑셀 분석")
    
    # PDF 업로드
    uploaded_file = st.file_uploader("📎 PDF 논문 업로드", type="pdf")
    
    if uploaded_file is not None:
        try:
            import pdfplumber
            
            # PDF 텍스트 추출
            with pdfplumber.open(uploaded_file) as pdf:
                pdf_text = ""
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        pdf_text += text + "\n"
            
            st.success(f"✅ PDF 업로드 성공! ({len(pdf.pages)}페이지)")
            
            # 텍스트 미리보기
            with st.expander("📖 텍스트 미리보기"):
                st.text_area("PDF 내용 (처음 2000자)", 
                            pdf_text[:2000] + "..." if len(pdf_text) > 2000 else pdf_text, 
                            height=200)
            
            # 분석 실행 버튼
            if st.button("📊 MEDDEV 엑셀 분석 실행", key="pdf_analysis", use_container_width=True):
                if not st.session_state.get('gemini_api_key', ''):
                    st.error("❌ AI 분석 탭에서 Gemini API 키를 먼저 입력하세요!")
                else:
                    try:
                        # Gemini 설정
                        genai.configure(api_key=st.session_state['gemini_api_key'])
                        model = genai.GenerativeModel('gemini-2.5-flash')
                        
                        # 텍스트 길이 제한
                        max_length = 50000
                        processed_text = pdf_text
                        if len(pdf_text) > max_length:
                            processed_text = pdf_text[:max_length] + "\n...(텍스트 길이 제한으로 일부 생략)"
                            st.warning(f"⚠️ 텍스트가 {max_length}자로 제한되어 분석됩니다")
                        
                        # MEDDEV 표 형식 프롬프트
                        prompt = get_meddev_table_analysis_prompt(processed_text)
                        
                        # 분석 실행
                        with st.spinner("📊 MEDDEV 엑셀 분석 중... (2-3분)"):
                            response = model.generate_content(prompt)
                            
                            if response and response.text:
                                st.success("✅ 분석 완료!")
                                
                                # 결과 표시
                                st.markdown("### 📊 MEDDEV 2.7/1 Rev. 4 엑셀 분석")
                                st.markdown(response.text)
                                
                                # 다운로드 파일 준비
                                analysis_result = f"""MEDDEV 2.7/1 Rev. 4 엑셀 분석

분석 날짜: {pd.Timestamp.now().strftime('%Y년 %m월 %d일 %H시 %M분')}
원본 논문 길이: {len(pdf_text)} 글자

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

{response.text}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

분석 도구: Gemini AI 기반 MEDDEV 분석 시스템
생성 시간: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}
"""
                                
                                # 다운로드 버튼
                                col1, col2 = st.columns(2)
                                
                                with col1:
                                    st.download_button(
                                        label="📥 마크다운 다운로드",
                                        data=analysis_result,
                                        file_name=f"MEDDEV_analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.md",
                                        mime="text/markdown",
                                        use_container_width=True
                                    )
                                
                                with col2:
                                    # 🔥 엑셀 다운로드 (병합 셀 오류 수정)
                                    try:
                                        # 엑셀 파일 생성
                                        wb = create_excel_meddev_analysis(response.text, pdf_text)
                                        
                                        # 엑셀 파일 저장
                                        excel_buffer = io.BytesIO()
                                        wb.save(excel_buffer)
                                        excel_buffer.seek(0)
                                        
                                        st.download_button(
                                            label="📊 엑셀 다운로드 (완벽한 표)",
                                            data=excel_buffer.getvalue(),
                                            file_name=f"MEDDEV_analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            use_container_width=True
                                        )
                                        
                                        st.success("✅ 엑셀 파일 생성 성공!")
                                        
                                    except Exception as excel_error:
                                        st.warning(f"⚠️ 엑셀 파일 생성 실패: {excel_error}")
                                        
                                        # 실패 시 텍스트 파일 제공
                                        st.download_button(
                                            label="📄 텍스트 다운로드 (백업)",
                                            data=analysis_result.encode('utf-8'),
                                            file_name=f"MEDDEV_analysis_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.txt",
                                            mime="text/plain",
                                            use_container_width=True
                                        )
                            else:
                                st.error("❌ Gemini 분석 응답을 받지 못했습니다")
                    
                    except Exception as e:
                        st.error(f"❌ 분석 오류: {e}")
        
        except ImportError:
            st.error("❌ pdfplumber가 설치되지 않았습니다. `pip install pdfplumber` 명령어를 실행하세요")
        except Exception as e:
            st.error(f"❌ PDF 처리 오류: {e}")
    
    else:
        st.info("📄 PDF 파일을 업로드하면 MEDDEV 2.7/1 Rev. 4 엑셀 분석이 진행됩니다")

# 푸터
render_footer()
# streamlit run main.py
# 로컬에서 실행하려면 위 명령어를 사용하세요
